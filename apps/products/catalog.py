import os
import xlsxwriter
from openpyxl.utils import get_column_letter
from rest_framework import generics
from django.http import HttpResponse
from .models import Products


class DownloadProductsXLS(generics.GenericAPIView):
    def get_image_filename(self, image_filename):
        return os.path.basename(image_filename) if image_filename else ""

    def get(self, request, *args, **kwargs):
        queryset = Products.objects.all()

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="products.xlsx"'

        wb = xlsxwriter.Workbook(response, {"in_memory": True})
        ws = wb.add_worksheet()

        columns = ["ID", "Category", "Title", "Price", "Image URL"]
        col_widths = [5, 40, 65, 8, 100]

        cell_format = wb.add_format({"color": "blue", "underline": 1})

        for col_idx, (column, width) in enumerate(zip(columns, col_widths), start=1):
            col_letter = get_column_letter(col_idx)
            ws.set_column(f"{col_letter}:{col_letter}", width)
            ws.write(0, col_idx - 1, column, cell_format)

        row = 1
        base_url = ""
        for product in queryset:
            image_filename = self.get_image_filename(product.image.name)

            data = [
                product.id,
                product.category.name,
                product.title,
                product.price,
                image_filename,
            ]

            for col_idx, value in enumerate(data, start=1):
                ws.write(row, col_idx - 1, value)

                if col_idx == len(data):
                    if image_filename:
                        full_image_url = (
                            f"{base_url}/media/products/{image_filename}"
                        )
                        ws.write_url(row, col_idx - 1, full_image_url, cell_format)

            row += 1

        wb.close()

        return response
